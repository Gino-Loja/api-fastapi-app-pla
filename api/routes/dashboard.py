from datetime import date, datetime
import ftplib
import io
import os
from fastapi import APIRouter, Body, Depends, File, Form, Request, Response, HTTPException, UploadFile, status
from fastapi.encoders import jsonable_encoder
from typing import Any, List, Optional
from fastapi.responses import FileResponse
from sqlalchemy import String, alias, cast, extract, text
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from zoneinfo import ZoneInfo
from sqlmodel import SQLModel, and_, select , func
from model import Areas, Comentarios, Comentarios_Dto, areas_profesor, Asignaturas, Periodo, Planificacion_Profesor, Planificaciones, Profesores
from api.deps import SessionDep, sender_email
from sqlalchemy.orm import aliased
from ftplib import FTP
import io
from typing import Optional
from pytz import timezone as tz
from utils import render_email_template_info, send_email
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse
import io  # Importar io para trabajar con archivos en memoria


router = APIRouter()
class Total_profesores(SQLModel):
    #__tablename__ = "profesores"
  
    total: int


@router.get("/areas/count", response_description="Obtener el total de áreas")
async def get_total_areas(session: SessionDep) -> Any:
    try:
        statement = select(func.count(Areas.id).label("total_areas"))
        result = session.exec(statement).one()
        return {"total_areas": result}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener el total de áreas"
        ) from e


@router.get("/profesores/count", response_description="Obtener el total de profesores", response_model=Total_profesores)
async def get_total_professors(session: SessionDep) -> Any:
    statement = select(func.count().label("total")).select_from(Profesores)
    result = session.exec(statement).one()
    return  {"total": result}


@router.get("/asignaturas/count", response_description="Obtener el total de asignaturas")
async def get_total_asignaturas(session: SessionDep) -> Any:
    try:
        # Consulta para contar la cantidad total de asignaturas
        statement = select(func.count(Asignaturas.id).label("total_asignaturas"))
        result = session.exec(statement).one()
        print(result)
        
        return {"total_asignaturas": result}

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener el total de asignaturas"
        ) from e
    
    
@router.get("/planificaciones/count-by-estado", response_description="Obtener la cantidad de planificaciones por estado")
async def get_planificaciones_count_by_estado(
    periodo_id: int,  # Nuevo parámetro para filtrar por período
    session: SessionDep
) :
    try:
        # Consulta para contar planificaciones agrupadas por estado y filtrar por período
        statement = (
            select(Planificacion_Profesor.estado, func.count(Planificacion_Profesor.id).label("total"))
            .join(Planificaciones, Planificaciones.id == Planificacion_Profesor.planificacion_id)  # Unir con Planificaciones
            .where(Planificaciones.periodo_id == periodo_id)  # Filtrar por período
            .group_by(Planificacion_Profesor.estado)
        )
        results = session.exec(statement).all()

        # Formatear los resultados como un diccionario
        estado_counts = {result.estado: result.total for result in results}

        return estado_counts
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener las planificaciones por estado"
        ) from e

@router.get("/docentes/atrasados", response_description="Obtener lista de docentes con planificaciones atrasadas")
async def get_docentes_atrasados(session: SessionDep, periodo_id: int) -> Any:
    try:
        # Consulta para obtener docentes con planificaciones atrasadas
        statement = (
            select(
                Profesores.id,
                Profesores.nombre,
                Planificaciones.titulo,
                Planificaciones.fecha_subida,
                Areas.nombre.label("area_nombre"),
                Asignaturas.nombre.label("asignatura_nombre"),
            )
            .join(Planificaciones, Planificaciones.profesor_id == Profesores.id)  # Unir Profesores con Planificaciones
            .join(Planificacion_Profesor, Planificacion_Profesor.planificacion_id == Planificaciones.id)  # Unir Planificaciones con Planificacion_Profesor
            .join(Asignaturas, Asignaturas.id == Planificaciones.asignaturas_id)  # Unir Planificaciones con Asignaturas
            .join(Areas, Areas.id == Asignaturas.area_id)  # Unir Asignaturas con Areas
            .where(
                Planificacion_Profesor.estado == "atrasado",  # Filtrar por estado "atrasado"
                Planificaciones.periodo_id == periodo_id  # Filtrar por periodo_id
            )
        )

        results = session.exec(statement).all()

        # Formatear los resultados
        docentes_atrasados = [
            {
                "id_profesor": result.id,
                "nombre_profesor": result.nombre,
                "titulo_planificacion": result.titulo,
                "fecha_subida": result.fecha_subida,
                "area_nombre": result.area_nombre,
                "asignatura_nombre": result.asignatura_nombre,
            }
            for result in results
        ]

        return docentes_atrasados
    except Exception as e:
        print("Error en la consulta:", str(e))  # Imprimir el error para depuración
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener docentes con planificaciones atrasadas: {str(e)}"
        ) from e

@router.get("/docentes/por-estado", response_description="Obtener lista de docentes según el estado de la planificación")
async def get_docentes_por_estado(session: SessionDep, estado: Optional[str] = None) -> Any:
    """
    Endpoint para obtener la lista de docentes según el estado de sus planificaciones.

    Parámetros:
    - estado (opcional): Estado de la planificación (e.g., "aprobado", "pendiente", "revisado", etc.).
      Si no se especifica, devuelve todos los docentes con planificaciones en cualquier estado.

    Retorna:
    - Una lista de docentes con planificaciones que coinciden con el estado especificado o con cualquier estado.
    """
    try:
        # Alias de tablas para facilitar la consulta
        profesores_alias = aliased(Profesores)
        planificaciones_profesor_alias = aliased(Planificacion_Profesor)
        planificaciones_alias = aliased(Planificaciones)

        # Construir la consulta base
        statement = (
            select(
                profesores_alias.id,
                profesores_alias.nombre,
                planificaciones_alias.titulo,
                planificaciones_profesor_alias.estado,
                planificaciones_profesor_alias.fecha_de_actualizacion
            )
            .join(planificaciones_alias, planificaciones_alias.profesor_id == profesores_alias.id)
            .join(planificaciones_profesor_alias, planificaciones_profesor_alias.planificacion_id == planificaciones_alias.id)
        )

        # Agregar filtro de estado si se proporciona
        if estado:
            statement = statement.where(planificaciones_profesor_alias.estado == estado)

        # Ejecutar la consulta
        results = session.exec(statement).all()

        # Formatear los resultados
        docentes = [
            {
                "id_profesor": result.id,
                "nombre_profesor": result.nombre,
                "titulo_planificacion": result.titulo,
                "estado_planificacion": result.estado,
                "fecha_actualizacion": result.fecha_de_actualizacion,
            }
            for result in results
        ]

        return {"estado": estado if estado else "todos", "docentes": docentes}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener la lista de docentes por estado"
        ) from e



@router.get("/docentes/estado", response_description="Obtener lista de docentes por estado")
async def get_usuarios_por_estado(session: SessionDep, estado: Optional[str] = None) -> Any:
    """
    Endpoint para obtener la lista de usuarios según su estado.

    Parámetros:
    - estado: "activo" o "inactivo". Si no se pasa, se devuelven todos los usuarios.

    Retorna:
    - Una lista de usuarios filtrada por el estado especificado.
    """
    try:
        # Construir la consulta base
        statement = select(Profesores.id, Profesores.nombre, Profesores.email, Profesores.estado)

        # Filtrar por estado si se especifica
        if estado:
            if estado.lower() == "activo":
                statement = statement.where(Profesores.estado == True)
            elif estado.lower() == "inactivo":
                statement = statement.where(Profesores.estado == False)
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Estado inválido. Los valores permitidos son 'activo' o 'inactivo'."
                )

        # Ejecutar la consulta
        results = session.exec(statement).all()

        # Formatear los resultados
        usuarios = [
            {
                "id_usuario": result.id,
                "nombre": result.nombre,
                "email": result.email,
                "estado_docente": "activo" if result.estado else "inactivo",
            }
            for result in results
        ]

        return {"estado": estado or "todos", "usuarios": usuarios}

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener la lista de usuarios"
        ) from e


#
@router.get("/metricas/total-planificaciones-asignadas", response_description="Obtener el total de planificaciones asignadas")
async def get_total_planificaciones_asignadas(session: SessionDep, periodo_id: int) -> Any:
    try:
        statement = select(func.count()).select_from(Planificacion_Profesor).join(Planificaciones, Planificaciones.id == Planificacion_Profesor.planificacion_id).where(Planificaciones.periodo_id == periodo_id)
        
        total = session.exec(statement).one()
        return {"total_planificaciones_asignadas": total}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener el total de planificaciones asignadas"
        ) from e
  #      

@router.get("/metricas/planificaciones-por-area", response_description="Obtener planificaciones por área")
async def get_planificaciones_por_area(session: SessionDep, periodo_id: int) -> Any:
    try:
        statement = (
            select(Areas.nombre, func.count().label("total_planificaciones"))
            .join(Asignaturas, Areas.id == Asignaturas.area_id)
            .join(Planificaciones, Asignaturas.id == Planificaciones.asignaturas_id)
            .join(Periodo, Periodo.id == Planificaciones.periodo_id)

            .group_by(Areas.nombre)
            .where(Planificaciones.periodo_id == periodo_id)

            
        )
        results = session.exec(statement).all()

        # Convertir las filas a una lista de diccionarios
        planificaciones = [{"nombre": row[0], "total_planificaciones": row[1]} for row in results]

        return  planificaciones
    except Exception as e:
        print(e)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener las métricas de planificaciones por área ${str(e)}"
        ) from e

        
@router.get("/metricas/planificaciones-aprobadas-vs-pendientes", response_description="Obtener planificaciones aprobadas vs pendientes")
async def get_planificaciones_aprobadas_vs_pendientes(session: SessionDep) -> Any:
    try:
        statement = (
            select(Planificacion_Profesor.estado, func.count().label("total"))
            .where(Planificacion_Profesor.estado.in_(["aprobado", "pendiente"]))
            .group_by(Planificacion_Profesor.estado)
        )
        results = session.exec(statement).all()
        
        if  len(results) == 0:
            return  {"aprobado": 0, "pendiente": 0}
        
        return  {result.estado: result.total for result in results}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener las métricas de planificaciones aprobadas vs pendientes"
        ) from e
        
@router.get("/metricas/profesores-con-mas-planificaciones-atrasadas", response_description="Obtener profesores con más planificaciones atrasadas")
async def get_profesores_con_mas_planificaciones_atrasadas(session: SessionDep, periodo_id: int):
    try:
        statement = (
            select(Profesores.nombre, func.count().label("total_atrasadas"))
            .join(Planificaciones, Profesores.id == Planificaciones.profesor_id)
            .join(Periodo, Planificaciones.periodo_id == Periodo.id)
            .join(Planificacion_Profesor, Planificaciones.id == Planificacion_Profesor.planificacion_id)  # Asegura la relación
            .where(Planificacion_Profesor.estado == "atrasado", Planificaciones.periodo_id == periodo_id)
            
            .group_by(Profesores.nombre)
            .order_by(func.count().desc())
            
        )
        results = session.exec(statement).all()
        
       

        # Convertir los resultados en una lista de diccionarios
        profesores_atrasados = [{"nombre": row[0], "total_atrasadas": row[1]} for row in results]

        return  profesores_atrasados
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener los profesores con más planificaciones atrasadas"
        ) from e

@router.get("/metricas/planificaciones-por-periodo", response_description="Obtener planificaciones por periodo")
async def get_planificaciones_por_periodo(session: SessionDep) -> Any:
    try:
        statement = (
            select(Periodo.nombre, func.count().label("total_planificaciones"))
            .join(Planificaciones, Periodo.id == Planificaciones.periodo_id)
            .group_by(Periodo.nombre)
        )
        results = session.exec(statement).all()

        # Convertir a un formato serializable
        planificaciones_por_periodo = [
            {"nombre_periodo": row.nombre, "total_planificaciones": row.total_planificaciones}
            for row in results
        ]

        return {"planificaciones_por_periodo": planificaciones_por_periodo}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener las métricas de planificaciones por periodo"
        ) from e


@router.get("/metricas/planificaciones-por-estado-por-area", response_description="Obtener el total de planificaciones por estado para cada área, agrupado por fecha_subida")
async def get_planificaciones_por_estado_por_area(session: SessionDep, periodo_id: int):
    try:
        # Consulta para obtener el total de planificaciones por estado para cada área, agrupado por fecha_subida
        statement = (
            select(
                Areas.nombre.label("nombre_area"),
                Planificacion_Profesor.estado,
                Planificaciones.fecha_subida,
                func.count().label("total_planificaciones")
            )
            .join(Asignaturas, Areas.id == Asignaturas.area_id)
            .join(Planificaciones, Asignaturas.id == Planificaciones.asignaturas_id)
            .join(Planificacion_Profesor, Planificaciones.id == Planificacion_Profesor.planificacion_id)
            .where(Planificaciones.periodo_id == periodo_id)
            .group_by(Areas.nombre, Planificacion_Profesor.estado, Planificaciones.fecha_subida)
        )

        results = session.exec(statement).all()

        # Formatear los resultados en una lista de diccionarios
        planificaciones_por_estado_por_area = [
            {
                "nombre_area": result.nombre_area,
                "estado": result.estado,
                "fecha_subida": result.fecha_subida,
                "total_planificaciones": result.total_planificaciones,
            }
            for result in results
        ]

        return planificaciones_por_estado_por_area

    except Exception as e:
        print(f"Error en la consulta: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener las métricas de planificaciones por estado por área: {str(e)}"
        ) from e

@router.get("/metricas/documentos-entregados-rango/", response_description="Listar documentos entregados en un rango de fechas", response_model=List[Any])
async def get_documentos_entregados_rango(
    fecha_inicio: date,  # Fecha de inicio del rango
    fecha_fin: date,     # Fecha de fin del rango
    session: SessionDep
) -> Any:
    try:
        # Crear alias para la tabla Profesores
        ProfesoresAsignado = aliased(Profesores)
        ProfesoresRevisor = aliased(Profesores)

        # Realizamos el join entre las tablas necesarias
        query = (
            select(
                Planificaciones.titulo,
                Planificaciones.descripcion,
                Planificaciones.fecha_subida,
                Planificacion_Profesor.fecha_de_actualizacion,
                Planificacion_Profesor.estado,
                ProfesoresAsignado.nombre.label("profesor_asignado_nombre"),  # Profesor asignado para subir
                areas_profesor.profesor_id.label("profesor_revisor_id"),  # ID del profesor revisor
                ProfesoresRevisor.nombre.label("profesor_revisor_nombre"),  # Nombre del profesor revisor
                Asignaturas.nombre.label("nombre_asignatura"),  # Nombre de la asignatura
                Asignaturas.curso,  # Curso de la asignatura
                Areas.nombre.label("nombre_area")  # Nombre del área
            )
            .join(Planificacion_Profesor, Planificaciones.id == Planificacion_Profesor.planificacion_id)
            .join(ProfesoresAsignado, Planificaciones.profesor_id == ProfesoresAsignado.id)  # Profesor asignado
            .join(areas_profesor, Planificacion_Profesor.profesor_revisor_id == areas_profesor.id, isouter=True)  # Profesor revisor
            .join(ProfesoresRevisor, areas_profesor.profesor_id == ProfesoresRevisor.id, isouter=True)  # Nombre del profesor revisor
            .join(Asignaturas, Planificaciones.asignaturas_id == Asignaturas.id)  # Asignatura
            .join(Areas, Asignaturas.area_id == Areas.id)  # Área
            .where(
                and_(
                    Planificacion_Profesor.fecha_de_actualizacion >= fecha_inicio,
                    Planificacion_Profesor.fecha_de_actualizacion <= fecha_fin
                )
            )  # Filtrar por rango de fechas
        )

        # Ejecutar la consulta y obtener los resultados
        documentos = session.exec(query).all()

        # Formatear la respuesta
        result = [
            {
                "titulo": documento.titulo,
                "descripcion": documento.descripcion,
                "fecha_subida": documento.fecha_subida,
                "fecha_actualizacion": documento.fecha_de_actualizacion,
                "estado": documento.estado,
                "profesor_asignado_nombre": documento.profesor_asignado_nombre,
                "profesor_revisor_nombre": documento.profesor_revisor_nombre,
                "nombre_asignatura": documento.nombre_asignatura,
                "curso": documento.curso,
                "nombre_area": documento.nombre_area,
            }
            for documento in documentos
        ]

        return result

    except Exception as e:
        print(f"Error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener los documentos entregados: {str(e)}"
        )








#mis planificaciones por estado
@router.get("/metricas/mis-planificaciones-por-estado/{profesor_id}", response_description="Obtener las planificaciones de un profesor por estado")
async def get_mis_planificaciones_por_estado(
    session: SessionDep, 
    profesor_id: int, 
    periodo_id: int  # Nuevo parámetro: periodo_id
) :
    try:
        # Consulta para obtener las planificaciones del profesor filtradas por periodo_id
        statement = (
            select(Planificacion_Profesor.estado, func.count().label("total"))
            .join(Planificaciones, Planificacion_Profesor.planificacion_id == Planificaciones.id)
            .where(
                Planificaciones.profesor_id == profesor_id,
                Planificaciones.periodo_id == periodo_id  # Filtro por periodo_id
            )
            .group_by(Planificacion_Profesor.estado)
        )
        results = session.exec(statement).all()

        # Formatear los resultados como un diccionario {estado: total}
        return {"mis_planificaciones_por_estado": {result.estado: result.total for result in results}}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener las métricas de planificaciones por estado"
        ) from e
        
        
@router.get("/metricas/mis-planificaciones-atrasadas/{profesor_id}", response_description="Obtener las planificaciones atrasadas de un profesor")
async def get_mis_planificaciones_atrasadas(session: SessionDep,
                                            profesor_id: int,
                                            periodo_id: int  # Nuevo parámetro: periodo_id
) :
    try:
        statement = (
            select(func.count())
            .select_from(Planificaciones)
            .join(Periodo, Planificaciones.periodo_id == Periodo.id)
            .where(Planificaciones.profesor_id == profesor_id,
                   Planificaciones.fecha_subida > Periodo.fecha_fin,
                   Planificaciones.periodo_id == periodo_id)
        )
        total = session.exec(statement).one()
        return {"mis_planificaciones_atrasadas": total}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener el total de planificaciones atrasadas"
        ) from e
        
@router.get("/metricas/mis-planificaciones-proximas-a-vencer/{profesor_id}", response_description="Obtener las planificaciones próximas a vencer de un profesor")
async def get_mis_planificaciones_proximas_a_vencer(session: SessionDep, profesor_id: int,periodo_id: int):
    try:
        # Consulta corregida
        statement = (
            select(Planificaciones.titulo, Planificaciones.fecha_subida, Asignaturas.nombre, Asignaturas.codigo)
            .join(Asignaturas, Planificaciones.asignaturas_id == Asignaturas.id)
            .join(Periodo, Planificaciones.periodo_id == Periodo.id)
            .join(Planificacion_Profesor, Planificaciones.id == Planificacion_Profesor.planificacion_id)
            .where(
                Planificaciones.profesor_id == profesor_id,  # Ajuste en la relación
                Planificaciones.periodo_id == periodo_id,
                
                Planificaciones.fecha_subida.between(
                    func.now(), func.now() + text("INTERVAL '7 days'")
                ),
                Planificacion_Profesor.estado == "pendiente"
            )
        )
        results = session.exec(statement).all()

        # Convertir los resultados a una lista de diccionarios
        planificaciones = [
            {
            "titulo": row.titulo,
            "fecha_subida": row.fecha_subida,
            "asignatura": row.nombre,
            "codigo": row.codigo,
            
            } for row in results
        ]

        return {"mis_planificaciones_proximas_a_vencer": planificaciones}
    except Exception as e:
        print(f"Error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener las planificaciones próximas a vencer"
        ) from e

        
@router.get("/planificaciones-asignadas/{profesor_id}", response_description="Obtener las planificaciones asignadas a un profesor con áreas y asignaturas")
async def get_planificaciones_asignadas(session: SessionDep, profesor_id: int, periodo_id: int) :
    try:
        # Consulta para obtener las planificaciones asignadas al profesor, junto con las áreas y asignaturas
        statement = (
            select(
                Planificaciones.id.label("id_planificacion"),
                Planificaciones.titulo,
                Planificaciones.descripcion,
                Planificaciones.fecha_subida,
                Planificacion_Profesor.estado,
                Asignaturas.nombre.label("nombre_asignatura"),
                Areas.nombre.label("nombre_area")
            )
            .join(Asignaturas, Planificaciones.asignaturas_id == Asignaturas.id)
            .join(Periodo, Planificaciones.periodo_id == Periodo.id)
            .join(Areas, Asignaturas.area_id == Areas.id)
            .join(Planificacion_Profesor, Planificaciones.id == Planificacion_Profesor.planificacion_id)
            .where(Planificaciones.profesor_id == profesor_id, Planificaciones.periodo_id == periodo_id)
        )

        results = session.exec(statement).all()

        # Formatear los resultados en una lista de diccionarios
        planificaciones = [
            {
                "id_planificacion": result.id_planificacion,
                "titulo": result.titulo,
                "descripcion": result.descripcion,
                "fecha_subida": result.fecha_subida,
                "estado": result.estado,
                "nombre_asignatura": result.nombre_asignatura,
                "nombre_area": result.nombre_area,
            }
            for result in results
        ]

        return planificaciones

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener las planificaciones asignadas"
        ) from e
        
        
@router.get("/asignaturas-con-planificaciones/{profesor_id}/{periodo_id}", response_description="Obtener las asignaturas con el número de planificaciones asignadas a un profesor en un periodo específico")
async def get_asignaturas_con_planificaciones(
    session: SessionDep, 
    profesor_id: int, 
    periodo_id: int
) :
    try:
        # Consulta para obtener las asignaturas con el número de planificaciones asignadas al profesor en el periodo especificado
        statement = (
            select(
                Asignaturas.id.label("id_asignatura"),
                Asignaturas.nombre.label("nombre_asignatura"),
                func.count(Planificaciones.id).label("total_planificaciones")
            )
            .join(Planificaciones, Asignaturas.id == Planificaciones.asignaturas_id)
            .join(Planificacion_Profesor, Planificaciones.id == Planificacion_Profesor.planificacion_id)
            .where(
                Planificaciones.profesor_id == profesor_id,
                Planificaciones.periodo_id == periodo_id
            )
            .group_by(Asignaturas.id, Asignaturas.nombre, Asignaturas.codigo)
        )

        results = session.exec(statement).all()

        # Formatear los resultados en una lista de diccionarios
        asignaturas = [
            {
                "id_asignatura": result.id_asignatura,
                "nombre_asignatura": result.nombre_asignatura,
                "total_planificaciones": result.total_planificaciones,
            }
            for result in results
        ]

        return asignaturas

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al obtener las asignaturas con planificaciones asignadas"
        ) from e
        
        

@router.get("/search/", response_description="Listar todas las planificaciones", response_model=List[Any])
async def get_all_planificaciones_descargar(
    periodo_id: int,  # Solo necesitamos el ID del periodo
    session: SessionDep
) -> Any:
    try:
        # Realizamos el join entre todas las tablas necesarias
        query = (
            select(
                Planificaciones,
                Planificacion_Profesor.id,
                Planificacion_Profesor.planificacion_id,
                Profesores.nombre.label("profesor_nombre"),
                Asignaturas.nombre.label("asignatura_nombre"),
                Asignaturas.curso.label("curso_nombre"),  # Agregar el nombre del curso
                Asignaturas.paralelo.label("paralelo"),  # Agregar el paralelo
                Periodo.nombre.label("periodo_nombre"),
                # Información del profesor aprobador
                Planificacion_Profesor.profesor_aprobador_id,
                # Información del profesor revisor
                Planificacion_Profesor.profesor_revisor_id,
                # Información del área
                Areas.id.label("area_id"),
                Areas.nombre.label("area_nombre"),
                Areas.codigo.label("area_codigo"),
                # Información adicional de la planificación
                Planificacion_Profesor.fecha_de_actualizacion,
                Planificacion_Profesor.estado,
                Planificacion_Profesor.archivo,
            )
            .join(Profesores, Planificaciones.profesor_id == Profesores.id)
            .join(Asignaturas, Planificaciones.asignaturas_id == Asignaturas.id)
            .join(Periodo, Planificaciones.periodo_id == Periodo.id)
            .join(Areas, Asignaturas.area_id == Areas.id)
            .join(
                Planificacion_Profesor,
                Planificaciones.id == Planificacion_Profesor.planificacion_id
            )
            # Join para obtener la información del profesor revisor
            .join(
                areas_profesor,
                Planificacion_Profesor.profesor_revisor_id == areas_profesor.id,
                isouter=True
            )
            .where(Planificaciones.periodo_id == periodo_id)  # Solo filtramos por periodo_id
        )
        
        # Ejecutar la consulta y obtener los resultados
        planificaciones = session.exec(query).all()

        # Obtener los IDs de profesores aprobadores y revisores
        profesor_aprobador_ids = [int(p[8]) for p in planificaciones if p[8] is not None]  # Convertir a entero
        profesor_revisor_ids = [int(p[9]) for p in planificaciones if p[9] is not None]    # Convertir a entero

        # Consultar nombres de profesores aprobadores
        aprobadores = {}
        if profesor_aprobador_ids:
            query_aprobadores = select(Profesores).where(Profesores.id.in_(profesor_aprobador_ids))
            aprobadores = {p.id: p.nombre for p in session.exec(query_aprobadores)}

        # Consultar nombres de profesores revisores a través de Areas_Profesor
        revisores = {}
        if profesor_revisor_ids:
            query_revisores = (
                select(areas_profesor, Profesores.nombre)
                .join(Profesores, areas_profesor.profesor_id == Profesores.id)
                .where(areas_profesor.id.in_(profesor_revisor_ids))
            )
            revisores = {ap.id: nombre for ap, nombre in session.exec(query_revisores)}

        # Crear una lista de resultados con los campos deseados
        result = [
            {
                # Información básica de la planificación
                "titulo": planificacion.titulo,
                "descripcion": planificacion.descripcion,
                "fecha_subida": planificacion.fecha_subida,
                
                # IDs de relaciones
                "profesor_id": planificacion.profesor_id,
                "asignaturas_id": planificacion.asignaturas_id,
                "periodo_id": planificacion.periodo_id,

                "id": id,
                "id_planificacion": planificacion_id,
                
                # Nombres de las relaciones principales
                "profesor_nombre": profesor_nombre,
                "periodo_nombre": periodo_nombre,
                "asignatura_nombre": asignatura_nombre,
                "curso_nombre": curso_nombre,  # Agregar el nombre del curso
                "paralelo": paralelo,  # Agregar el paralelo
                
                # Información del área
                "area_id": area_id,
                "area_nombre": area_nombre,
                "area_codigo": area_codigo,
                
                # Información del profesor aprobador
                "profesor_aprobador_id": profesor_aprobador_id,
                "profesor_aprobador_nombre": aprobadores.get(profesor_aprobador_id) if profesor_aprobador_id else None,
                
                # Información del profesor revisor
                "profesor_revisor_id": profesor_revisor_id,
                "profesor_revisor_nombre": revisores.get(profesor_revisor_id) if profesor_revisor_id else None,
                
                # Información adicional de la planificación
                "fecha_de_actualizacion": fecha_de_actualizacion,
                "estado": estado,
                "archivo": archivo,
            }
            for (
                planificacion, id, planificacion_id, profesor_nombre, asignatura_nombre, curso_nombre, paralelo, periodo_nombre,
                profesor_aprobador_id, profesor_revisor_id,
                area_id, area_nombre, area_codigo,
                fecha_de_actualizacion, estado, archivo
            ) in planificaciones
        ]
        
        return result

    except Exception as e:
        print(f"Error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener las planificaciones: {str(e)}"
        )






# Mapa de colores para los estados
status_color_map = {
    'entregado': '22C55E',  # verde
    'pendiente': 'EAB308',  # amarillo
    'atrasado': 'EF4444',   # rojo
    'aprobado': '3B82F6',   # azul
    'revisado': 'A855F7'    # morado
}

@router.get("/download-planificaciones-excel/", response_description="Descargar planificaciones en formato Excel")
async def download_planificaciones_excel(
    periodo_id: int,  # ID del periodo
    session: SessionDep
) -> Any:
    try:
        # Obtener los datos del endpoint existente
        planificaciones = await get_all_planificaciones_descargar(periodo_id, session)

        # Crear un nuevo libro de Excel en memoria
        wb = Workbook()
        ws = wb.active
        ws.title = "Planificaciones"

        # Definir los encabezados de las columnas
        headers = [
            "Título", "Descripción", "Fecha de Subida", "Profesor", "Asignatura", "Curso", "Paralelo", "Periodo",
            "Área", "Profesor Aprobador", "Profesor Revisor", "Fecha de Actualización", "Estado", "Archivo"
        ]
        ws.append(headers)

        # Llenar el archivo Excel con los datos
        for planificacion in planificaciones:
            row = [
                planificacion["titulo"],
                planificacion["descripcion"],
                planificacion["fecha_subida"].strftime("%Y-%m-%d %H:%M:%S"),
                planificacion["profesor_nombre"],
                planificacion["asignatura_nombre"],
                planificacion["curso_nombre"],
                planificacion["paralelo"],  # Incluir el paralelo
                planificacion["periodo_nombre"],
                planificacion["area_nombre"],
                planificacion["profesor_aprobador_nombre"],
                planificacion["profesor_revisor_nombre"],
                planificacion["fecha_de_actualizacion"].strftime("%Y-%m-%d %H:%M:%S"),
                planificacion["estado"],
                planificacion["archivo"]
            ]
            ws.append(row)

            # Aplicar el color de fondo solo a la celda de la columna "Estado"
            estado = planificacion["estado"].lower()
            if estado in status_color_map:
                fill = PatternFill(start_color=status_color_map[estado], end_color=status_color_map[estado], fill_type="solid")
                # La columna "Estado" es la columna 12 (índice 11 en base 0)
                ws.cell(row=ws.max_row, column=13).fill = fill  # Columna 13 es "Estado"

        # Crear un archivo en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)  # Mover el puntero al inicio del archivo

        # Nombre del archivo
        filename = f"planificaciones_periodo_{periodo_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Devolver el archivo como una respuesta de streaming
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"Error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el archivo Excel: {str(e)}"
        )